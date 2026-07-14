#!/usr/bin/env python3.12
"""Ekstrak kurikulum dari Program Semester Excel + PDF SHOFI"""
import openpyxl

# ── 1. Parse Excel ──
excel_files = {
    "Matematika Penalaran": "moodle-files/MTK_Penalaran.xlsx",
    "Matematika": "moodle-files/Matematika_Wajib.xlsx",
}

for subject, fname in excel_files.items():
    wb = openpyxl.load_workbook(fname)
    print(f"\n{'='*60}")
    print(f"=== {subject} ===")
    print('='*60)
    for sn in wb.sheetnames:
        ws = wb[sn]
        semester = "Ganjil" if "ganjil" in sn else "Genap"
        print(f"\n-- Semester {semester} --")
        # Cari header row dulu (kolom No, Elemen, ATP, Konsep, Waktu)
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            vals = [str(v).strip().lower()[:20] if v else '' for v in row[:5]]
            if 'no' in vals and 'elemen' in vals and 'waktu' in vals:
                header_row = i
                break
        
        if not header_row:
            print("  (header not found)")
            continue
        
        # Baca data mulai dari header_row + 1
        elemen = ""
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            no = str(row[0]).strip() if row[0] else ''
            konsep = str(row[3]).strip() if row[3] else ''
            jam_str = str(row[4]).strip() if row[4] else ''
            
            # Skip baris kosong, header, total, penilaian
            if not konsep or not jam_str:
                continue
            if konsep.lower().startswith('total') or konsep.lower().startswith('penilaian') or konsep.lower().startswith('remedial') or konsep.lower().startswith('penguatan'):
                continue
            
            try:
                jam = int(float(jam_str.replace(',', '.')))
                if jam <= 0:
                    continue
                print(f"  [wk?] {konsep[:80]} ({jam} jam)")
            except (ValueError, TypeError):
                pass

# ── 2. PDF files need OCR ──
print("\n\n=== PDF files (need OCR) ===")
import subprocess
pdf_files = [
    ("B. Indonesia Ganjil", "moodle-files/B_Indonesia_Ganjil.pdf"),
    ("B. Indonesia Genap", "moodle-files/B_Indonesia_Genap.pdf"),
    ("Biologi", "moodle-files/Biologi_Promes.pdf"),
]
for label, path in pdf_files:
    print(f"\n{label} ({path}):")
    # Try pdftotext first
    import os
    txt_path = path + ".txt"
    result = subprocess.run(["pdftotext", "-nopgbrk", path, txt_path], capture_output=True, text=True)
    if result.returncode == 0 and os.path.exists(txt_path):
        with open(txt_path) as f:
            content = f.read()
        print(f"  Length: {len(content)} chars")
        # Show first 500 chars
        print(f"  Preview: {content[:500]}")
    else:
        print(f"  pdftotext error: {result.stderr[:200]}")
