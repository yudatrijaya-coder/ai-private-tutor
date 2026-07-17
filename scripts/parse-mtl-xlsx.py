import openpyxl
import json

wb = openpyxl.load_workbook("public/moodle-files/Program_Semester_2627_MTL_XI.xlsx", data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ===")
    i = 0
    for row in ws.iter_rows(values_only=True):
        i += 1
        clean = [str(c) if c is not None else "" for c in row]
        clean_str = " | ".join(clean)
        if any(c.strip() for c in clean):
            print(f"  R{i}: {clean_str}")
